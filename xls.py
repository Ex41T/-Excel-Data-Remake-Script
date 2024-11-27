import pandas as pd
import openpyxl
from colorama import Fore
import re
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl import load_workbook


packets = {
        "Pakiet Klasycznywaga: 1.15kg - 150 g. w plastrach (vacum)": [
        {"name": "Krakowska Waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Polędwica  Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Szynka ChudaWaga:150 g /  0.5 kg / 1 kg /  - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Baleron Waga: 150g / 0.5kg / 1 kg  - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Boczek Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Kiełbasa Śląska waga: ok. 400g", "quantity": 1},
    ],
    "Pakiet Klasycznywaga: 1.15kg - 150 g. w plastrach (papier)": [
        {"name": "Krakowska Waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Polędwica  Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Szynka ChudaWaga:150 g /  0.5 kg / 1 kg /  - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Baleron Waga: 150g / 0.5kg / 1 kg  - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Boczek Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Kiełbasa Śląska waga: ok. 400g", "quantity": 1},
    ],
    "Pakiet Klasyczny XL<i>waga:</i>1.9 kg - 150 g. w plastrach (papier)": [
        {"name": "Krakowska Waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Polędwica  Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Szynka ChudaWaga:150 g /  0.5 kg / 1 kg /  - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Baleron Waga: 150g / 0.5kg / 1 kg  - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Boczek Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Kiełbasa Śląska waga: ok. 400g", "quantity": 1},
    ],
    "Pakiet Klasyczny XL<i>waga:</i>1.9 kg - 150 g. w plastrach (vacum)": [
        {"name": "Krakowska Waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 2},
        {"name": "Polędwica  Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 2},
        {"name": "Szynka ChudaWaga:150 g /  0.5 kg / 1 kg /  - 150 g. w plastrach (vacum)", "quantity": 2},
        {"name": "Baleron Waga: 150g / 0.5kg / 1 kg  - 150 g. w plastrach (vacum)", "quantity": 2},
        {"name": "Boczek Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 2},
        {"name": "Kiełbasa Śląska waga: ok. 400g", "quantity": 1},
    ],
    "Pakiet Kiełbaswaga: 2kg": [
        {"name": "Kiełbasa Jałowcowa waga: ok. 450g", "quantity": 1},
        {"name": "Kiełbasa Śląska waga: ok. 400g", "quantity": 1},
        {"name": "Kiełbasa Swojskawaga: ok. 350g", "quantity": 1},
        {"name": "Kiełbasa Myśliwska waga: ok. 450g", "quantity": 1},
        {"name": "Serdelki  waga: ok. 450g", "quantity": 1},
    ],
    "Pakiet Wiejski<i>waga: 1.75kg - 150 g. w plastrach (vacum)": [
        {"name": "Bekon Waga:150g / 0.5kg / 1kg - 150 g. w plastrach (vacum)", "quantity": 2},
        {"name": "Polędwica Łososiowa waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 2},
        {"name": "Szynka z Tłuszczem Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Polędwiczka Wieprzowa Waga:ok. 400g", "quantity": 1},
        {"name": "Kiełbasa Myśliwska waga: ok. 450g", "quantity": 1},
    ],
     "Pakiet Wiejski<i>waga: 1.75kg - 150 g. w plastrach (papier)": [
        {"name": "Bekon Waga:150g / 0.5kg / 1kg - 0.5 kg w plastrach (papier)", "quantity": 2},
        {"name": "Polędwica Łososiowa waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Szynka z Tłuszczem Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Polędwiczka Wieprzowa Waga:ok. 400g", "quantity": 1},
        {"name": "Kiełbasa Myśliwska waga: ok. 450g", "quantity": 1},
    ],
    "Pakiet Wędlin waga: 1.65kg - 150 g. w plastrach (papier)": [
        {"name": "Baleron Waga: 150g / 0.5kg / 1 kg  - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Bekon Waga:150g / 0.5kg / 1kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Boczek Surowy waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Boczek Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Krakowska Waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Polędwica  Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Polędwica Łososiowa waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Salceson Waga:300 g - 300g(papier)", "quantity": 1},
        {"name": "Szynka ChudaWaga:150 g /  0.5 kg / 1 kg /  - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Szynka z Tłuszczem Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
    ],
    "Pakiet Wędlin waga: 1.65kg - 150 g. w plastrach (vacum)": [
        {"name": "Baleron Waga: 150g / 0.5kg / 1 kg  - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Bekon Waga:150g / 0.5kg / 1kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Boczek Surowy waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Boczek Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Krakowska Waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Polędwica  Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Polędwica Łososiowa waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Salceson Waga:300 g - 300g(vacum)", "quantity": 1},
        {"name": "Szynka ChudaWaga:150 g /  0.5 kg / 1 kg /  - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Szynka z Tłuszczem Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
    ],
    "Pakiet Pasztetów waga: 1.2kg": [
        {"name": "Pasztet z Śliwkąwaga:400g", "quantity": 1},
        {"name": "Pasztet z Żurawinąwaga:400g", "quantity": 1},
        {"name": "Pasztetwaga:400g", "quantity": 1},
    ],
}


def clean_html(content):
    return re.sub(r'<br>|<I>|<\/I>|<b>|<\/b>', '', content)


def process_packets(df):

    """

    Breaks down packets into individual items with their quantities.

    """

    additional_rows = []
    for index, row in df.iterrows():
        item_name = row['Item Name']
        quantity = int(row['Quantity (- Refund)'])
        if item_name in packets:
            for content in packets[item_name]:
                for _ in range(quantity):
                    full_item_name = content["name"]
                    additional_rows.append([full_item_name, content["quantity"]])
        else:
            additional_rows.append([item_name, quantity])
    
    new_df = pd.DataFrame(additional_rows, columns=['Item Name', 'Quantity (- Refund)'])
    grouped_df = new_df.groupby('Item Name').sum().reset_index()
    
    return grouped_df



def move_customer_note_to_end(df_kurierzy):

    if "Customer Note" in df_kurierzy.columns:
        cols = [col for col in df_kurierzy.columns if col != "Customer Note"] + ["Customer Note"]
        df_kurierzy = df_kurierzy[cols]
    return df_kurierzy


def highlight_phrase(worksheet, phrase="w kawałku"):
   
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and phrase in cell.value:
                cell.fill = red_fill


def extract_product_name(item_name):

    """

    Extracts the product name from a string up to the first occurrence of 'Weight' or similar terms.

    """

    match = re.match(r"(.+?)\s*(Waga|waga)", item_name)
    if match:
        return match.group(1).strip()
    return item_name



EXCLUDE_PRODUCTS = [
    "Eko Twaróg Półtłusty 220g Prosto z Farmy",
    "Masło Extra 82% Tłuszczu Eko 200 G", 
    "Eko Jogurt Naturalny  250g",
    "Smalec ze Skwarkami", 
    "Eko Mleko 3,9 %  750ml", 
    "Eko Zsiadłe Mleko 450g", 
    "Eko Śmietana 18% 250g", 
    "Jajka Ekologiczne  ilość: 10 szt. ", 
    "Polędwiczka Wieprzowa", 
    "Pasztet z Śliwką", 
    "Pasztet z Żurawiną", 
    "Pasztet",
    "Kiełbasa Myśliwska",
    "Kiełbasa Jałowcowa",
    "Kiełbasa Swojska",
    "Pakiet Pasztetów"
]


# Func to exclude products
def should_exclude_product(product_name):
    return product_name in EXCLUDE_PRODUCTS


def calculate_weight_based_on_type(item_name, quantity_sum):

    """

    Calculates total weight based on product type.

    """

    if "w plastrach" in item_name:
        return 0.150 * quantity_sum 
    elif "w kawałku" in item_name:
        if "0.5 kg" in item_name:
            return 0.500 * quantity_sum  
        elif "1 kg" in item_name:
            return 1.000 * quantity_sum  
    return 0



def calc_total_weight(df):
    PRODUCT_WEIGHTS = {
    "Kiełbasa Śląska": 0.400,  
    "Biała Kiełbasa": 0.500,   
    "Serdelki": 0.450,         
    "Kaszanka": 0.450,         
    "Salceson": 0.300          
}

    products_weights = {}

    for index, row in df.iterrows():
        product_name = extract_product_name(row['Item Name'])  
        quantity = row['Quantity (- Refund)']

       
        if should_exclude_product(product_name):
            continue

        # Get weight of the product
        base_weight = PRODUCT_WEIGHTS.get(product_name)
        
        if base_weight:
            
            total_weight = base_weight * quantity
        else:
            # Calc weight for other products e.g in pieces or slices
            total_weight = calculate_weight_based_on_type(row['Item Name'], quantity)

        # Sum qunatities for the same product (repeated cells with the same content)
        if product_name in products_weights:
            products_weights[product_name] += total_weight
        else:
            products_weights[product_name] = total_weight

    # convert do dataframe
    total_weight_df = pd.DataFrame(list(products_weights.items()), columns=["Produkt", "Suma(kg)"])

    return total_weight_df


#sheet styles
def adjust_column_widths(worksheet, df):
    for i, column in enumerate(df):
        max_length = max(df[column].astype(str).map(len).max(), len(str(column))) + 2
        worksheet.column_dimensions[get_column_letter(i+1)].width = max_length


def make_columns_bold(df, worksheet, columns):
    bold_font = Font(bold=True)
    for col_name in columns:
        col_idx = get_column_letter(df.columns.get_loc(col_name) + 1)
        for row in range(1, worksheet.max_row + 1):
            worksheet[col_idx + str(row)].font = bold_font



def save_total_weights_to_excel(input_file_path, df_total_weights):
    with pd.ExcelWriter(input_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_total_weights.to_excel(writer, sheet_name='Produkty Gramatury', index=False)


def open_file_dialog():
    Tk().withdraw()
    return askopenfilename()


def save_all_data_to_excel(input_file_path, df_produkty, df_total_weights, df_kurierzy):
    with pd.ExcelWriter(input_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_produkty.to_excel(writer, sheet_name='Produkty', index=False)
        df_total_weights.to_excel(writer, sheet_name='Produkty Gramatury', index=False)
        df_kurierzy.to_excel(writer, sheet_name='Kurierzy', index=False)


def main():
    input_file_path = open_file_dialog()
    if not input_file_path:
        print("File selection was canceled.")
        return

    # Load the original Excel file
    df_original = pd.read_excel(input_file_path)
    df_processed = df_original.copy()
    df_processed['Item Name'] = df_processed['Item Name'].apply(clean_html)
    
    #Packet processing and repeat 'items name' column content calc
    df_produkty = process_packets(df_processed)

    # calc weight
    df_total_weights = calc_total_weight(df_produkty)

    # Convert raw data for 'Kurierzy' sheet
    columns_to_delete = [
        'Line number', 'Email (Billing)', 'First Name (Shipping)',
        'Last Name (Shipping)', 'Address 1&2 (Shipping)', 'City (Shipping)', 'Postcode (Shipping)',
        'Order Subtotal Amount', 'Order Shipping Amount', 'SKU', 'Item #',
        'Item Name', 'Quantity (- Refund)'
    ]
    df_kurierzy = df_original.drop(columns=[col for col in columns_to_delete if col in df_original.columns], errors='ignore')
    df_kurierzy = df_kurierzy.drop_duplicates(subset=['Order Number'], keep='first')
    df_kurierzy.reset_index(drop=True, inplace=True)
    df_kurierzy = move_customer_note_to_end(df_kurierzy)


    save_all_data_to_excel(input_file_path, df_produkty, df_total_weights, df_kurierzy)

    # Reopen workbook, adjust styles for both sheets
    workbook = openpyxl.load_workbook(input_file_path)
    worksheet_produkty = workbook['Produkty']
    highlight_phrase(worksheet_produkty, phrase="w kawałku")
    adjust_column_widths(worksheet_produkty, df_produkty)
    make_columns_bold(df_produkty, worksheet_produkty, ['Item Name', 'Quantity (- Refund)'])
    
    worksheet_kurierzy = workbook['Kurierzy']
    adjust_column_widths(worksheet_kurierzy, df_kurierzy)

    workbook.save(input_file_path)

    print(f"File saved >>> {input_file_path}")


if __name__ == "__main__":
    main()
