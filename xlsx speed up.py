import pandas as pd
import openpyxl
from colorama import Fore
import re
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl import load_workbook

packets = {
        "Pakiet Klasycznywaga: 1.15kg - 150 g. w plastrach (vacum)": [
        {"name": "Krakowska Waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Polędwica  Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Szynka ChudaWaga:150 g /  0.5 kg / 1 kg /  - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Baleron Waga: 150g / 0.5kg / 1 kg  - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Boczek Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Kiełbasa Śląska waga: ok. 400g", "quantity": 1},
    ],
    "Pakiet Klasycznywaga: 1.15kg - 150 g. w plastrach (papier)": [
        {"name": "Krakowska Waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Polędwica  Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Szynka ChudaWaga:150 g /  0.5 kg / 1 kg /  - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Baleron Waga: 150g / 0.5kg / 1 kg  - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Boczek Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Kiełbasa Śląska waga: ok. 400g", "quantity": 1},
    ],
    "Pakiet Klasyczny XL<i>waga:</i>1.9 kg - 150 g. w plastrach (papier)": [
        {"name": "Krakowska Waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Polędwica  Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Szynka ChudaWaga:150 g /  0.5 kg / 1 kg /  - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Baleron Waga: 150g / 0.5kg / 1 kg  - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Boczek Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Kiełbasa Śląska waga: ok. 400g", "quantity": 1},
    ],
    "Pakiet Klasyczny XL<i>waga:</i>1.9 kg - 150 g. w plastrach (vacum)": [
        {"name": "Krakowska Waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 2},
        {"name": "Polędwica  Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 2},
        {"name": "Szynka ChudaWaga:150 g /  0.5 kg / 1 kg /  - 150 g. w plastrach (vacum)", "quantity": 2},
        {"name": "Baleron Waga: 150g / 0.5kg / 1 kg  - 150 g. w plastrach (vacum)", "quantity": 2},
        {"name": "Boczek Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 2},
        {"name": "Kiełbasa Śląska waga: ok. 400g", "quantity": 1},
    ],
    "Pakiet Kiełbaswaga: 2kg": [
        {"name": "Kiełbasa Jałowcowa waga: ok. 450g", "quantity": 1},
        {"name": "Kiełbasa Śląska waga: ok. 400g", "quantity": 1},
        {"name": "Kiełbasa Swojskawaga: ok. 350g", "quantity": 1},
        {"name": "Kiełbasa Myśliwska waga: ok. 450g", "quantity": 1},
        {"name": "Serdelki  waga: ok. 450g", "quantity": 1},
    ],
    "Pakiet Wiejski<i>waga: 1.75kg - 150 g. w plastrach (vacum)": [
        {"name": "Bekon Waga:150g / 0.5kg / 1kg - 150 g. w plastrach (vacum)", "quantity": 2},
        {"name": "Polędwica Łososiowa waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 2},
        {"name": "Szynka z Tłuszczem Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Polędwiczka Wieprzowa Waga:ok. 400g", "quantity": 1},
        {"name": "Kiełbasa Myśliwska waga: ok. 450g", "quantity": 1},
    ],
     "Pakiet Wiejski<i>waga: 1.75kg - 150 g. w plastrach (papier)": [
        {"name": "Bekon Waga:150g / 0.5kg / 1kg - 0.5 kg w plastrach (papier)", "quantity": 2},
        {"name": "Polędwica Łososiowa waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Szynka z Tłuszczem Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 2},
        {"name": "Polędwiczka Wieprzowa Waga:ok. 400g", "quantity": 1},
        {"name": "Kiełbasa Myśliwska waga: ok. 450g", "quantity": 1},
    ],
    "Pakiet Wędlin waga: 1.65kg - 150 g. w plastrach (papier)": [
        {"name": "Baleron Waga: 150g / 0.5kg / 1 kg  - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Bekon Waga:150g / 0.5kg / 1kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Boczek Surowy waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Boczek Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Krakowska Waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Polędwica  Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Polędwica Łososiowa waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Salceson Waga:300 g - 300g(papier)", "quantity": 1},
        {"name": "Szynka ChudaWaga:150 g /  0.5 kg / 1 kg /  - 150 g. w plastrach (papier)", "quantity": 1},
        {"name": "Szynka z Tłuszczem Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (papier)", "quantity": 1},
    ],
    "Pakiet Wędlin waga: 1.65kg - 150 g. w plastrach (vacum)": [
        {"name": "Baleron Waga: 150g / 0.5kg / 1 kg  - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Bekon Waga:150g / 0.5kg / 1kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Boczek Surowy waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Boczek Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Krakowska Waga:150g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Polędwica  Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Polędwica Łososiowa waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Salceson Waga:300 g - 300g(vacum)", "quantity": 1},
        {"name": "Szynka ChudaWaga:150 g /  0.5 kg / 1 kg /  - 150 g. w plastrach (vacum)", "quantity": 1},
        {"name": "Szynka z Tłuszczem Waga:150 g / 0.5 kg / 1 kg - 150 g. w plastrach (vacum)", "quantity": 1},
    ],
    "Pakiet Pasztetów waga: 0.9kg": [
        {"name": "Pasztet z Śliwkąwaga:300g", "quantity": 1},
        {"name": "Pasztet z Żurawinąwaga:300g", "quantity": 1},
        {"name": "Pasztetwaga:300g", "quantity": 1},
    ],
}
#oczyszczenie kontentu html
def clean_html(content):
    return re.sub(r'<br>|<I>|<\/I>|<b>|<\/b>', '', content)

#Przetworzenie pakietów 
def process_packets(df):
    additional_rows = []
    for index, row in df.iterrows():
        item_name = row['Item Name']
        quantity = int(row['Quantity (- Refund)'])  #ilość jest liczbą całkowitą do podsumowania
        if item_name in packets:
            for content in packets[item_name]:
                for _ in range(quantity):
                    # Dodajemy opcje takie jak "papier" czy "vacum" do produktu
                    full_item_name = content["name"]
                    additional_rows.append([full_item_name, content["quantity"]])
        else:
            additional_rows.append([item_name, quantity])
    
    # Konwertacja do DataFrame'u
    new_df = pd.DataFrame(additional_rows, columns=['Item Name', 'Quantity (- Refund)'])
    
    # Pogrupuj kolumne 'Item Name' i zrób sume w kolumnie 'Quantity (- Refund)', uwzględniając różne opcje
    grouped_df = new_df.groupby('Item Name').sum().reset_index()
    
    return grouped_df

def move_customer_note_to_end(df_kurierzy):
    # Czy istnieje kolumna Customer Note
    if "Customer Note" in df_kurierzy.columns:
        # Przenosimy kolumnę "Customer Note" na sam koniec
        customer_note = df_kurierzy.pop("Customer Note")  # Wyciągamy kolumnę
        df_kurierzy["Customer Note"] = customer_note      # Wstawiamy ją na koniec
    return df_kurierzy

def highlight_phrase(worksheet, phrase="w kawałku"):
    # ozaczenie kawałków - czerwone tło
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # Iterujemy przez wszystkie komórki w arkuszu
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and phrase in cell.value:
                # Jeżeli fraza "w kawałku" jest w komórce, ustawiamy czerwone tło
                cell.fill = red_fill

#stylowanie kolumn
def adjust_column_widths(worksheet, df):
    for i, column in enumerate(df):
        max_length = max(df[column].astype(str).map(len).max(), len(str(column))) + 2
        worksheet.column_dimensions[get_column_letter(i+1)].width = max_length

def make_columns_bold(df, worksheet, columns):
    bold_font = Font(bold=True)
    for col_name in columns:
        col_idx = get_column_letter(df.columns.get_loc(col_name) + 1)
        for row in range(1, worksheet.max_row + 1):
            worksheet[col_idx + str(row)].font = bold_font


# Lista produktów do wykluczenia
EXCLUDE_PRODUCTS = [
    "Eko Jogurt", "Eko Mleko", "Eko Zsiadłe Mleko", "Eko Śmietana", "Jajka Ekologiczne", 
    "Polędwiczka Wieprzowa", "Pasztet z Śliwką", "Pasztet z Żurawiną", "Pasztet","Kiełbasa Myśliwska","Kiełbasa Jałowcowa","Kiełbasa Swojska"
]

# Wyciąganie nazwy produktu do słowa "Waga" lub spacji przed "Waga"
def extract_product_name(item_name):
    match = re.match(r"(.+?)\s*(Waga|waga)", item_name)
    if match:
        return match.group(1).strip()
    return item_name

# Funkcja określająca wagę dla produktów w plastrach lub kawałkach
def calculate_weight_based_on_type(item_name, quantity_sum):
    if "w plastrach" in item_name:
        return 0.150 * quantity_sum  # 150g = 0.150kg
    elif "w kawałku" in item_name:
        return 0.500 * quantity_sum  # 500g = 0.500kg
    return 0

# Funkcja do wykluczania nieistotnych produktów
def should_exclude_product(product_name):
    for keyword in EXCLUDE_PRODUCTS:
        if keyword in product_name:
            return True
    return False

# Funkcja do obliczania sumarycznej wagi dla wędlin i kiełbas
def calculate_total_weight(df):
    products_weights = {}
    
    for index, row in df.iterrows():
        product_name = extract_product_name(row['Item Name'])  # Wyciągamy nazwę produktu
        quantity = row['Quantity (- Refund)']
        
        # Sprawdzamy, czy produkt należy wykluczyć
        if should_exclude_product(product_name):
            continue
        
        # Obliczamy wagę dla kiełbas (na podstawie ilości i stałej wagi np. 450g)
        if "Kiełbasa Śląska" in product_name:
            total_weight = 0.400 * quantity  # 400g na sztukę
        elif "Biała Kiełbasa" in product_name:
            total_weight = 0.500 * quantity  # 500g na sztukę
        elif "Serdelki" in product_name:
            total_weight = 0.450 * quantity  # 450g na sztukę
        elif "Kaszanka" in product_name:
            total_weight = 0.450 * quantity  # 450g na sztukę
        elif "Salceson" in product_name:
            total_weight = 0.300 * quantity  # Zakładamy, że 300g na sztukę
        else:
            # Obliczamy dla innych produktów: plastrach lub kawałkach
            total_weight = calculate_weight_based_on_type(row['Item Name'], quantity)
        
        # Sumujemy ilości dla tego samego produktu
        if product_name in products_weights:
            products_weights[product_name] += total_weight
        else:
            products_weights[product_name] = total_weight
    
    # Konwertujemy wyniki do DataFrame
    total_weight_df = pd.DataFrame(list(products_weights.items()), columns=["Produkt", "Suma(kg)"])
    
    return total_weight_df

# Funkcja zapisywania wyników do nowego arkusza Excela
def save_total_weights_to_excel(input_file_path, df_total_weights):
    with pd.ExcelWriter(input_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_total_weights.to_excel(writer, sheet_name='Produkty Gramatury', index=False)

def open_file_dialog():
    Tk().withdraw()
    return askopenfilename()

def main():
    input_file_path = open_file_dialog()
    if not input_file_path:
        print("File selection was canceled.")
        return

    # Load the original Excel file
    df_original = pd.read_excel(input_file_path)
    df_processed = df_original.copy()
    df_processed['Item Name'] = df_processed['Item Name'].apply(clean_html)
    
    # Przetwarzanie pakietów i obliczanie powtórzonych produktów
    df_produkty = process_packets(df_processed)

  # Obliczanie wagi całkowitej i tworzenie nowego sheet'u
    df_total_weights = calculate_total_weight(df_produkty)
    save_total_weights_to_excel(input_file_path, df_total_weights)

    # Zapisywanie info do "Produkty" sheet
    with pd.ExcelWriter(input_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_produkty.to_excel(writer, sheet_name='Produkty', index=False)

    # Przetwarzanie informacji do "Kurierzy" sheet
    columns_to_delete = [
        'Line number', 'Email (Billing)', 'First Name (Shipping)',
        'Last Name (Shipping)', 'Address 1&2 (Shipping)', 'City (Shipping)', 'Postcode (Shipping)',
        'Order Subtotal Amount', 'Order Shipping Amount', 'SKU', 'Item #',
        'Item Name', 'Quantity (- Refund)'
    ]
    df_kurierzy = df_original.drop(columns=[col for col in columns_to_delete if col in df_original.columns], errors='ignore')
    df_kurierzy = df_kurierzy.drop_duplicates(subset=['Order Number'], keep='first')
    df_kurierzy.reset_index(drop=True, inplace=True)
    df_kurierzy = move_customer_note_to_end(df_kurierzy)

    # Zapisywanie informacji do "Kurierzy" sheet
    with pd.ExcelWriter(input_file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df_kurierzy.to_excel(writer, sheet_name='Kurierzy', index=False)

    # Otwórz ponownie skoroszyt, aby dostosować style obu arkuszy
    workbook = openpyxl.load_workbook(input_file_path)
    worksheet_produkty = workbook['Produkty']
    highlight_phrase(worksheet_produkty, phrase="w kawałku")
    adjust_column_widths(worksheet_produkty, df_produkty)
    make_columns_bold(df_produkty, worksheet_produkty, ['Item Name', 'Quantity (- Refund)'])
    
    worksheet_kurierzy = workbook['Kurierzy']
    adjust_column_widths(worksheet_kurierzy, df_kurierzy)

    
    workbook.save(input_file_path)

    print(f"Zapisane >>> {input_file_path}")

if __name__ == "__main__":
    main()  
